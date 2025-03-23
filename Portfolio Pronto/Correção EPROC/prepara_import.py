def prepara_import():
    import datetime
    import pandas as pd
    from openpyxl import load_workbook
    import shutil

    data_atual = datetime.date.today().strftime("%d%m%y")

    # Copiando arquivo base a partir do modelo
    shutil.copy('Modelo EPROC ATT.xlsx', rf'.\SmartImports\Correcao Digit EPROC ATT - {data_atual}.xlsx')

    # Importando e limpando planilha de desdobramentos
    desdobramentos = pd.read_excel('relatorio_desdobramentos.xlsx')
    colunas_tabela = list(desdobramentos.columns)
    colunas_novas = list(desdobramentos.iloc[2])
    dici = {}
    for x in range(len(colunas_tabela)):
        dici[colunas_tabela[x]] = colunas_novas[x]
    desdobramentos.rename(columns=dici, inplace=True)
    desdobramentos.drop(axis=0, index=[0, 1, 2], inplace=True)

    # Contrói dataframe da lista de notas
    df_notas = pd.read_excel('Lista de notas.xlsx')

    # Filtra os casos de "Migrado" e "Digitalizado" e armazena em uma lista, assim como os erros
    resultado_linha = []
    erros = []
    for linha in df_notas.iterrows():
        if linha[1]['Status 1'] == 'Migrado' or linha[1]['Status 1'] == 'Digitalizado':
            insercao = [linha[1]['Processo'], linha[1]['originario_1'], linha[1]['Status 1']]
            resultado_linha.append(insercao)
        elif linha[1]['Status 2'] == 'Migrado' or linha[1]['Status 2'] == 'Digitalizado':
            insercao = [linha[1]['Processo'], linha[1]['originario_2'], linha[1]['Status 2']]
            resultado_linha.append(insercao)
        elif linha[1]['Status 2'] == 'Migrado' or linha[1]['Status 3'] == 'Digitalizado':
            insercao = [linha[1]['Processo'], linha[1]['originario_3'], linha[1]['Status 3']]
            resultado_linha.append(insercao)
        else:
            erros.append(linha[1])

    # cria e exporta dataframe do resultado filtrado
    df_filtrado = pd.DataFrame(resultado_linha, columns=['Processo', 'Originario', 'Status'])
    df_filtrado.to_excel('Resultado filtrado.xlsx', index=False)
    print(df_filtrado)
    print('Processos fora das hipóteses de digitalização:', len(erros))

    # Descobre a pasta dos processos a partir do processo originário
    resultado_final = []
    for item in df_filtrado.iterrows():
        if item[1]['Originario'] in list(desdobramentos['numero']) and item[1]['Processo'] not in list(desdobramentos['numero']):
            index = list(desdobramentos['numero']).index(item[1]['Originario'])
            pasta_desdobramento = desdobramentos.iloc[index, 2]
            print(pasta_desdobramento)
            resultado_final.append([pasta_desdobramento[:9],
                                    pasta_desdobramento,
                                    item[1]['Originario'],
                                    item[1]['Processo']])
        else:
            print(item[1]['Originario'], '- não encontrado')

    # Cria um dataframe com os novos dados
    df_resultado_final = pd.DataFrame(resultado_final,
                                      columns=['Pasta', 'Pasta desdobramento', 'Número antigo', 'Processo'])
    print(df_resultado_final.head())

    # Carrega a planilha modelo base
    modelo_att = load_workbook(rf'.\SmartImports\Correcao Digit EPROC ATT - {data_atual}.xlsx')
    worksheet = modelo_att.active

    # Seleciona a planilha na qual será inserido o novo dataframe
    writer = pd.ExcelWriter(rf'.\SmartImports\Correcao Digit EPROC ATT - {data_atual}.xlsx', mode='a', if_sheet_exists='overlay')

    # Adiciona o novo dataframe na planilha existente
    df_resultado_final.to_excel(writer, index=False, header=False, sheet_name='Importacao', startrow=worksheet.max_row)

    # Salva as mudanças na planilha
    writer.close()

    return


prepara_import()

